'''
Микросервис на архитектуре REST API для
загрузки файла в формате *.xlsx на сервер,
импорта данных из файла в Базу Данных и
выдачи данных из SQL-таблицы в формате JSON.

'''

import os

import pandas as pd
import psycopg2
from dateutil import parser
from flask import Flask, jsonify, redirect, request
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from sqlalchemy import MetaData, Table, create_engine, select
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = './download'
ALLOWED_EXTENSIONS = {'xlsx'}
app = Flask(__name__)
t_host = 'localhost'
t_port = '5432'
t_dbname = 'db'
t_user = 'postgres'
t_pw = '1ZakonOma'
app.config['SQLALCHEMY_DATABASE_URI'
           ] = f'postgresql://{t_user}:{t_pw}@{t_host}/{t_dbname}'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
db = SQLAlchemy(app)
migrate = Migrate(app, db)
metadata = MetaData()
engine = create_engine(f'postgresql://{t_user}:{t_pw}@{t_host}/{t_dbname}')
db_conn = psycopg2.connect(host=t_host, port=t_port,
                           dbname=t_dbname, user=t_user, password=t_pw)


class DataModel(db.Model):
    '''
    Метод создания модели данных.
    '''
    __tablename__ = 'data'

    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.DateTime())
    delta = db.Column(db.Float())

    def __init__(self, date, delta):
        self.date = date
        self.delta = delta

    def __repr__(self):
        return f"{self.id}"


def allowed_file(filename):
    '''
    Метод проверки расширения файла.
    '''
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    '''
    Метод загрузки файла на сервер.
    '''
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            return redirect(request.url)
        if (file and allowed_file(file.filename)
                and (file.filename == 'testData.xlsx')):
            filename = secure_filename(file.filename)
            path = './download/testData.xlsx'
            os.remove(path)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            return download_file('Файл успешно загружен')
    return '''
    <!doctype html>
    <title>Загрузка вашего файла</title>
    <style>
        body {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            min-height: 100vh;
            color: #f0f8ff;
            background-color: #070217;
            font-size: 24px;
            font-family: Helvetica, sans-serif;
        }
        </style>
    <body>
    <h1>Загрузка вашего файла</h1>
    <form method=post enctype=multipart/form-data>
      <input type=file name=file>
      <input type=submit value=Загрузить>
    </form>
    </body>
    '''


@app.route('/download/', methods=['GET'])
def download_file(message):
    '''
    Метод оповещения об успешной загрузке файла.
    '''
    return '''
    <!doctype html>
    <title>%s</title>
    <style>
        body {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            min-height: 100vh;
            color: #f0f8ff;
            background-color: #070217;
            font-size: 24px;
            font-family: Helvetica, sans-serif;
        }
        </style>
    <body>
    <h1>%s</h1>
    </body>
    ''' % (message, message)


@app.route('/import/xlsx/', methods=['GET'])
def import_file():
    '''
    Метод загрузки данных из файла .xlsx
    в базу данных.
    '''
    dataset = DataModel.query.filter_by().all()
    if dataset:
        for object in dataset:
            db.session.delete(object)
            db.session.commit()
        return jsonify(
            result=0,
            message='Данные из БД удалены'
        )
    wb = load_workbook('./download/testData.xlsx')
    sheet = wb.get_sheet_by_name('Лист1')
    amount = sheet.max_row
    for counter in range(2, amount+1):
        date_string = sheet.cell(row=counter, column=1).value
        datetime_object = parser.parse(date_string)
        delta = float(
            str(sheet.cell(row=counter, column=2).value).replace(',', '.'))
        data = DataModel(datetime_object, delta)
        db.session.add(data)
        db.session.commit()
    return jsonify(
        result=amount-1,
        message='Данные введены в БД'
    )


@app.route('/export/sql/', methods=['GET'])
def export_json():
    '''
    Метод вывода данных из Представления
    SQL в формате JSON.
    '''
    create_view()
    dataset = Table('dataview', metadata, autoload_with=engine)
    stmt = select(dataset)
    connection = engine.connect()
    dataset = connection.execute(stmt).fetchall()
    if dataset:
        date_list = {}
        delta_list = {}
        deltalag_list = {}
        lag_num_list = {}
        counter = 1
        for object in dataset:
            date = object.date.strftime('%d.%m.%Y')
            delta = object.delta
            deltalag = 0 if object.deltalag is None else object.deltalag
            lag_num = abs(object.delta - (
                float(0 if object.deltalag is None else object.deltalag)))
            date_list[counter] = date
            delta_list[counter] = delta
            deltalag_list[counter] = deltalag
            lag_num_list[counter] = lag_num
            counter += 1
        return jsonify(
            message='OK',
            Rep_dt=date_list,
            delta=delta_list,
            deltalag=deltalag_list,
            lag_num=lag_num_list,
        )
    return jsonify(
        message='Данные отсутствуют',
        result=0
    )


@ app.route('/export/pandas/', methods=['GET'])
def export_pandas():
    '''
    Метод вывода данных из БД в формате JSON
    с помощью библиотеки Pandas.
    '''
    dataset = DataModel.query.filter_by().all()
    if dataset:
        date_list = []
        delta_list = []
        for object in dataset:
            date_list.append(object.date)
            delta_list.append(object.delta)
        df = pd.DataFrame({'Rep_dt': date_list,
                           'delta': delta_list})
        df = df.sort_values(by='Rep_dt')
        df['DeltaLag'] = (df['delta'].shift(2, fill_value=0))
        df['lag_num'] = (df['delta']-df['DeltaLag']).abs()
        return df.to_json(path_or_buf=None, date_format='iso')
    return jsonify(
        data=0,
        message='Данные отсутствуют'
    )


@ app.route('/create/', methods=['GET'])
def create_view():
    """
    Метод создания VIEW Представления
    в PostgreSQL.
    """
    db_cursor = db_conn.cursor()
    s = '''
    CREATE OR REPLACE VIEW dataview AS
    WITH result_data AS (
         SELECT data.date,
            data.delta
           FROM data
          ORDER BY data.date
        )
    SELECT result_data.date,
    result_data.delta,
    lag(result_data.delta, 2) OVER
    (ORDER BY result_data.date) AS deltalag
    FROM result_data;
    '''
    try:
        db_cursor.execute(s)
        db_conn.commit()
    except psycopg2.Error as e:
        e = str(e)
        return jsonify(t_message=e)
    db_cursor.close()
    return jsonify(t_message="Представление VIEW создано")


if __name__ == '__main__':
    app.run(debug=True)
